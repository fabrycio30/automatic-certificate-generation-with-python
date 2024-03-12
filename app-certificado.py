from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Spacer, Image, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from datetime import datetime

#automatização de certificados com python
from openpyxl import load_workbook
from docx import Document


# Função para criar o certificado
def create_certificate(nome_aluno, cpf_aluno, curso, carga_horaria):
    doc = SimpleDocTemplate(f'./todos/certificado_{nome_aluno.replace(" ", "_")}.pdf', pagesize=letter)

    title_style = ParagraphStyle(
        name='Title',
        parent=getSampleStyleSheet()['Title'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.black,
        alignment=1,
        spaceAfter=20
    )

    body_style = ParagraphStyle(
        name='BodyText',
        parent=getSampleStyleSheet()['Normal'],
        fontName='Helvetica',
        fontSize=12,
        textColor=colors.black,
        alignment=1,
        spaceAfter=20
    )

    #  logo
    logo = "./logo.png"  
    im = Image(logo, width=70, height=90)


    
    # Texto do certificado
    today = datetime.now().strftime("%d/%m/%y")
    content = [
        im,
        Spacer(1, 5),
        Paragraph(f'<b>CERTIFICADO</b>', title_style),
        Spacer(1, 12),
        Paragraph(f"Este certificado atesta que <b>{nome_aluno}</b>, portador do CPF <b>{cpf_aluno}</b>, concluiu com êxito o curso de <b>{curso}</b>, totalizando <b>{int(carga_horaria)}</b> horas de dedicação e aprendizado.", body_style),
        Spacer(1, 20),
        Paragraph(f"Dado o exposto, concedemos este certificado como reconhecimento oficial de que <b>{nome_aluno}</b> completou com sucesso o curso, atestando sua competência e comprometimento.", body_style),
        Spacer(1, 20),
        Paragraph(f"<b>Data de Emissão:</b> {today}", body_style),
        Spacer(1, 50),
        Paragraph(f"<b>Responsável:</b>  <i>Fabrício A. do Carmo</i>", body_style),
        Spacer(1, 20),
    ]

    doc.build(content)


planilha_fornecedores = load_workbook('./alunos.xlsx')
pagina_da_planilha = planilha_fornecedores['certificados']

for amostra in pagina_da_planilha.iter_rows(min_row=2, values_only=True):
    #print(amostra)
    nome_aluno, cpf_aluno, curso, carga_horaria = amostra
    # Chamar a função para criar o certificado
    create_certificate(nome_aluno, cpf_aluno, curso, carga_horaria)

